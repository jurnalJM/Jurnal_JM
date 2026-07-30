"""
Excel Export Utility
Export transaction data to Excel files
"""

from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import EXPORT_DIR


class ExcelExporter:
    """Export transactions to Excel files"""

    # Colors
    HEADER_FILL = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    ALTERNATE_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    CURRENCY_FORMAT = "#,##0.00"
    DATE_FORMAT = "dd-mmm-yy"

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    @staticmethod
    def export_transactions(
        transactions: List[Any],
        filename: str = None,
        include_summary: bool = True
    ) -> str:
        """
        Export transactions to Excel.

        Args:
            transactions: List of Transaksi objects
            filename: Optional custom filename
            include_summary: Include summary sheet

        Returns:
            Path to created file
        """
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Add transactions sheet
        ExcelExporter._add_transactions_sheet(wb, transactions)

        # Add summary sheet if requested
        if include_summary:
            ExcelExporter._add_summary_sheet(wb, transactions)

        # Generate filename if not provided
        if not filename:
            filename = f"Transaksi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Ensure export directory exists
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)

        # Save file
        filepath = EXPORT_DIR / filename
        wb.save(str(filepath))

        return str(filepath)

    @staticmethod
    def _add_transactions_sheet(wb: Workbook, transactions: List[Any]):
        """Add transactions data sheet"""
        ws = wb.create_sheet("Transaksi")

        # Headers
        headers = [
            "ID", "Tanggal", "Nota", "Dealer", "Nama Pembeli",
            "HP", "No Mesin", "Type Motor", "Warna",
            "DP", "Subsidi", "Diskon", "Insentif",
            "Leasing", "Tgl Lunas", "Pelunasan", "Status"
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = ExcelExporter.HEADER_FONT
            cell.fill = ExcelExporter.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = ExcelExporter.THIN_BORDER

        # Write data
        for row_num, trans in enumerate(transactions, 2):
            row_data = [
                trans.id,
                trans.tanggal,
                trans.nota,
                trans.dealer.nama if trans.dealer else "",
                trans.nama_pembeli,
                trans.telp_pembeli or "",
                trans.motor.no_mesin if trans.motor else "",
                trans.motor.type_motor.nama_type if trans.motor else "",
                trans.motor.warna if trans.motor else "",
                float(trans.detail.dp or 0) if trans.detail else 0,
                float(trans.detail.subsidi or 0) if trans.detail else 0,
                float((trans.detail.diskon or 0) + (trans.detail.diskon_tambahan or 0)) if trans.detail else 0,
                float(trans.detail.insentif or 0) if trans.detail else 0,
                trans.leasing.nama if trans.leasing else "",
                trans.detail.tgl_lunas if trans.detail else None,
                float(trans.detail.pelunasan or 0) if trans.detail else 0,
                ExcelExporter._get_status_text(trans.status_transaksi),
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = ExcelExporter.THIN_BORDER

                # Format based on column
                if col_num in [2, 15]:  # Dates
                    cell.number_format = ExcelExporter.DATE_FORMAT
                elif col_num in [10, 11, 12, 13, 16]:  # Currency
                    cell.number_format = ExcelExporter.CURRENCY_FORMAT
                    cell.alignment = Alignment(horizontal="right")

                # Alternate row colors
                if row_num % 2 == 0:
                    cell.fill = ExcelExporter.ALTERNATE_FILL

        # Auto-width columns
        for col_num, header in enumerate(headers, 1):
            max_length = len(header)
            column = get_column_letter(col_num)

            for row in ws.iter_rows(min_col=col_num, max_col=col_num, min_row=2):
                cell = row[0]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column].width = min(max_length + 2, 50)

    @staticmethod
    def _add_summary_sheet(wb: Workbook, transactions: List[Any]):
        """Add summary sheet"""
        ws = wb.create_sheet("Ringkasan", 0)

        # Calculate totals
        total_transactions = len(transactions)
        total_dp = sum(float(t.detail.dp or 0) for t in transactions if t.detail)
        total_subsidi = sum(float(t.detail.subsidi or 0) for t in transactions if t.detail)
        total_diskon = sum(
            float((t.detail.diskon or 0) + (t.detail.diskon_tambahan or 0))
            for t in transactions if t.detail
        )
        total_insentif = sum(float(t.detail.insentif or 0) for t in transactions if t.detail)
        total_pelunasan = sum(float(t.detail.pelunasan or 0) for t in transactions if t.detail)

        # Summary data
        summary_data = [
            ("Total Transaksi", total_transactions),
            ("Total DP", total_dp),
            ("Total Subsidi", total_subsidi),
            ("Total Diskon", total_diskon),
            ("Total Insentif", total_insentif),
            ("Total Pelunasan", total_pelunasan),
        ]

        # Write summary
        for row_num, (label, value) in enumerate(summary_data, 2):
            # Label
            label_cell = ws.cell(row=row_num, column=1)
            label_cell.value = label
            label_cell.font = Font(bold=True)
            label_cell.border = ExcelExporter.THIN_BORDER

            # Value
            value_cell = ws.cell(row=row_num, column=2)
            value_cell.value = value
            if isinstance(value, float):
                value_cell.number_format = ExcelExporter.CURRENCY_FORMAT
            value_cell.alignment = Alignment(horizontal="right")
            value_cell.border = ExcelExporter.THIN_BORDER

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20

        # Summary by dealer
        ExcelExporter._add_dealer_summary(wb, transactions)

    @staticmethod
    def _add_dealer_summary(wb: Workbook, transactions: List[Any]):
        """Add dealer breakdown sheet"""
        ws = wb.create_sheet("Per Dealer")

        # Group by dealer
        dealer_data = {}
        for trans in transactions:
            dealer_name = trans.dealer.nama if trans.dealer else "Unknown"
            if dealer_name not in dealer_data:
                dealer_data[dealer_name] = {
                    "count": 0,
                    "total_dp": 0,
                    "total_subsidi": 0,
                    "total_diskon": 0,
                }
            dealer_data[dealer_name]["count"] += 1
            if trans.detail:
                dealer_data[dealer_name]["total_dp"] += float(trans.detail.dp or 0)
                dealer_data[dealer_name]["total_subsidi"] += float(trans.detail.subsidi or 0)
                dealer_data[dealer_name]["total_diskon"] += float(
                    (trans.detail.diskon or 0) + (trans.detail.diskon_tambahan or 0)
                )

        # Headers
        headers = ["Dealer", "Jumlah", "Total DP", "Total Subsidi", "Total Diskon"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = ExcelExporter.HEADER_FONT
            cell.fill = ExcelExporter.HEADER_FILL
            cell.border = ExcelExporter.THIN_BORDER

        # Data
        for row_num, (dealer, data) in enumerate(sorted(dealer_data.items()), 2):
            row_data = [
                dealer,
                data["count"],
                data["total_dp"],
                data["total_subsidi"],
                data["total_diskon"],
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = ExcelExporter.THIN_BORDER

                if col_num > 1:
                    cell.number_format = ExcelExporter.CURRENCY_FORMAT if col_num > 2 else "0"
                    cell.alignment = Alignment(horizontal="right")

                if row_num % 2 == 0:
                    cell.fill = ExcelExporter.ALTERNATE_FILL

        # Auto-width
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 18

    @staticmethod
    def _get_status_text(status_code: str) -> str:
        """Convert status code to text"""
        status_map = {
            "P": "Pending",
            "A": "Approved",
            "L": "Lunas",
            "C": "Cancelled",
        }
        return status_map.get(status_code, status_code)
