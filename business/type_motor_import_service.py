"""
Import Service for Type Motor from Excel (Harga OTR format)
"""

from datetime import date, datetime
from typing import List, Dict, Any
from pathlib import Path
import openpyxl
import xlrd

from database.connection import DatabaseManager
from database.models import TypeMotor


class TypeMotorImportService:
    """Service for importing type motor from Excel files"""

    def __init__(self):
        self.errors = []
        self.warnings = []
        self.updated_rows = []

    def import_from_excel(self, file_path: str, tgl_expired: date = None) -> Dict[str, Any]:
        """
        Import type motor from Excel file (Harga OTR format)

        Args:
            file_path: Path to Excel file
            tgl_expired: Tanggal expired harga (if not provided, default 1 month from today)

        Returns:
            Dict with results, errors, warnings
        """
        self.errors = []
        self.warnings = []
        self.updated_rows = []

        # Default expired date: 1 month from today if not provided
        if tgl_expired is None:
            from dateutil.relativedelta import relativedelta
            tgl_expired = date.today() + relativedelta(months=1)

        try:
            # Determine file format and read accordingly
            file_path_obj = Path(file_path)
            is_old_format = file_path_obj.suffix.lower() == '.xls'

            rows_data = []

            if is_old_format:
                # Use xlrd for old .xls format
                wb = xlrd.open_workbook(file_path)
                ws = wb.sheet_by_index(0)

                for row_num in range(ws.nrows):
                    values = [ws.cell_value(row_num, col_idx) for col_idx in range(ws.ncols)]
                    if not any(values):
                        continue
                    rows_data.append((row_num + 1, values))
            else:
                # Use openpyxl for .xlsx format
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active

                for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
                    values = [cell.value for cell in row]
                    if not any(values):
                        continue
                    rows_data.append((row_num, values))

            # Parse and validate rows
            for row_num, values in rows_data:
                result = self._parse_row(row_num, values, tgl_expired)
                if result:
                    self.updated_rows.append(result)

            # Update database
            updated_count = 0
            if self.updated_rows and not self.errors:
                updated_count = self._update_database(self.updated_rows)

            return {
                'success': len(self.errors) == 0,
                'updated': updated_count,
                'total': len(self.updated_rows),
                'errors': self.errors,
                'warnings': self.warnings,
                'preview': self.updated_rows[:5],
            }

        except Exception as e:
            self.errors.append(f"Failed to read Excel file: {str(e)}")
            return {
                'success': False,
                'updated': 0,
                'total': 0,
                'errors': self.errors,
                'warnings': self.warnings,
            }

    def _parse_row(self, row_num: int, values: List, tgl_expired: date) -> Dict[str, Any]:
        """
        Parse and validate a single row from Excel (Harga OTR format)

        Excel columns:
        A: Tgl Efektif
        B: Kode Type
        C: Kode Motor
        D: Nama Type
        E: Nosin (prefix mesin/no. mesin)
        F: NoKa (prefix rangka/no. rangka)
        G: Harga OTR
        """
        # Skip header row
        if row_num == 1:
            return None

        if len(values) < 7:
            self.errors.append(f"Row {row_num}: Tidak cukup kolom (minimal 7)")
            return None

        kode_type = values[1]
        nama_type = values[3]
        nosin = values[4]  # E: No. Mesin prefix
        noka = values[5]   # F: No. Rangka prefix (might be formula)
        harga_otr = values[6]

        # Validate required fields
        if not kode_type:
            self.errors.append(f"Row {row_num}: Kode Type kosong")
            return None
        if not nama_type:
            self.errors.append(f"Row {row_num}: Nama Type kosong")
            return None
        if not nosin:
            self.errors.append(f"Row {row_num}: No. Mesin kosong")
            return None

        # Parse harga OTR
        try:
            otr = float(harga_otr) if harga_otr else 0
        except:
            otr = 0

        # Parse NoKa (might be formula like ="MH1"&E2, extract the prefix part)
        noka_str = str(noka).strip() if noka else ""

        # If NoKa contains formula, extract prefix (everything before &)
        if '&' in noka_str and '=' in noka_str:
            # Extract prefix from formula like ="MH1"&E2 → MH1
            try:
                parts = noka_str.split('&')[0]  # Get ="MH1"
                noka_prefix = parts.replace('=', '').replace('"', '').strip()
            except:
                noka_prefix = noka_str
        else:
            noka_prefix = noka_str

        if not noka_prefix:
            self.errors.append(f"Row {row_num}: No. Rangka kosong")
            return None

        return {
            'row_num': row_num,
            'kode_type': str(kode_type).strip(),
            'nama_type': str(nama_type).strip(),
            'prefix_nomesin': str(nosin).strip(),
            'prefix_norangka': noka_prefix,
            'otr': int(otr),
            'tgl_expired_harga': tgl_expired,
        }

    def _update_database(self, rows: List[Dict]) -> int:
        """Update or create type motor records in database"""
        session = DatabaseManager.get_session()
        updated = 0

        try:
            for row_data in rows:
                kode_type = row_data['kode_type']

                # Try to find existing type motor
                existing = session.query(TypeMotor).filter_by(kode_type=kode_type).first()

                if existing:
                    # Update existing
                    existing.nama_type = row_data['nama_type']
                    existing.prefix_nomesin = row_data['prefix_nomesin']
                    existing.prefix_norangka = row_data['prefix_norangka']
                    existing.tgl_expired_harga = row_data['tgl_expired_harga']
                    if row_data['otr'] > 0:
                        existing.otr = row_data['otr']
                    session.commit()
                    updated += 1
                else:
                    # Create new
                    new_type = TypeMotor(
                        kode_type=kode_type,
                        nama_type=row_data['nama_type'],
                        prefix_nomesin=row_data['prefix_nomesin'],
                        prefix_norangka=row_data['prefix_norangka'],
                        otr=row_data['otr'],
                        tgl_expired_harga=row_data['tgl_expired_harga'],
                        status='A',
                    )
                    session.add(new_type)
                    session.commit()
                    updated += 1
                    self.warnings.append(f"Type motor baru dibuat: {kode_type}")

        except Exception as e:
            session.rollback()
            self.errors.append(f"Database error: {str(e)}")
        finally:
            session.close()

        return updated
