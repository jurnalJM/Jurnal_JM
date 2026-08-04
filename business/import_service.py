"""
Import Service for Excel-based Stok Motor import
Handles parsing and validation of distribusi Excel files
"""

from datetime import date
from typing import List, Dict, Any, Tuple
import openpyxl
import xlrd
from pathlib import Path

from database.connection import DatabaseManager
from database.models import StokMotor, TypeMotor
from database.import_config import get_dealer_id
from business.exceptions import ValidationException


class ImportService:
    """Service for importing stok motor from Excel files"""

    def __init__(self):
        self.errors = []
        self.warnings = []
        self.validated_rows = []
        self.type_motor_suggestions = {}  # Suggestions for type motor format

    def import_from_excel(self, file_path: str, tgl_datang: date = None) -> Dict[str, Any]:
        """
        Import stok motor from Excel file

        Args:
            file_path: Path to Excel file
            tgl_datang: Default tanggal masuk (if not in Excel)

        Returns:
            Dict with results, errors, warnings
        """
        self.errors = []
        self.warnings = []
        self.validated_rows = []

        if tgl_datang is None:
            tgl_datang = date.today()

        try:
            # Determine file format and read accordingly
            file_path_obj = Path(file_path)
            is_old_format = file_path_obj.suffix.lower() == '.xls'

            rows_data = []

            if is_old_format:
                # Use xlrd for old .xls format
                wb = xlrd.open_workbook(file_path)
                ws = wb.sheet_by_index(0)

                for row_num in range(ws.nrows):
                    values = [ws.cell_value(row_num, col_idx) for col_idx in range(ws.ncols)]
                    # Skip empty rows
                    if not any(values):
                        continue
                    rows_data.append((row_num + 1, values))
            else:
                # Use openpyxl for .xlsx format
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active

                for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
                    values = [cell.value for cell in row]
                    # Skip empty rows
                    if not any(values):
                        continue
                    rows_data.append((row_num, values))

            # Parse and validate rows
            for row_num, values in rows_data:
                result = self._parse_row(row_num, values, tgl_datang)
                if result:
                    self.validated_rows.append(result)

            # Import validated rows
            imported_count = 0
            if self.validated_rows and not self.errors:
                imported_count = self._insert_to_database(self.validated_rows)

            # Enhance suggestions with type motor info
            suggestions_enhanced = []
            if self.type_motor_suggestions:
                suggestion_session = DatabaseManager.get_session()
                for type_id, suggestion in self.type_motor_suggestions.items():
                    type_motor = suggestion_session.query(TypeMotor).filter_by(id=type_id).first()
                    if type_motor:
                        suggestions_enhanced.append({
                            'type_id': type_id,
                            'type_kode': type_motor.kode_type,
                            'type_nama': type_motor.nama_type,
                            'prefix_norangka': suggestion['prefix_norangka'],
                            'prefix_nomesin': suggestion['prefix_nomesin'],
                        })
                suggestion_session.close()

            return {
                'success': len(self.errors) == 0,
                'imported': imported_count,
                'total': len(self.validated_rows),
                'errors': self.errors,
                'warnings': self.warnings,
                'preview': self.validated_rows[:5],  # First 5 rows for preview
                'type_motor_suggestions': suggestions_enhanced,  # Format suggestions for master type
            }

        except Exception as e:
            self.errors.append(f"Failed to read Excel file: {str(e)}")
            return {
                'success': False,
                'imported': 0,
                'total': 0,
                'errors': self.errors,
                'warnings': self.warnings,
            }

    def _parse_row(self, row_num: int, values: List, tgl_datang: date) -> Dict[str, Any]:
        """
        Parse and validate a single row from Excel (stokmotor.xls format)

        Excel columns:
        A: Tanggal Masuk (optional, uses default tgl_datang if empty)
        B: Dealer (nama dealer dari database)
        C: No. Mesin
        D: No. Rangka
        E: Kode Type (ML1F, MJ1E, dll)
        F: Warna (nama warna atau kode)
        G: Link (optional, unused)
        """
        # Skip header row
        if row_num == 1:
            return None

        if len(values) < 6:
            self.errors.append(f"Row {row_num}: Tidak cukup kolom (minimal 6)")
            return None

        tgl_masuk_val = values[0]
        dealer_nama = values[1]
        no_mesin = values[2]
        no_rangka = values[3]
        kode_type = values[4]
        warna = values[5]

        # Validate required fields
        if not no_mesin:
            self.errors.append(f"Row {row_num}: Nomor Mesin kosong")
            return None
        if not no_rangka:
            self.errors.append(f"Row {row_num}: Nomor Rangka kosong")
            return None
        if not kode_type:
            self.errors.append(f"Row {row_num}: Kode Type kosong")
            return None
        if not dealer_nama:
            self.errors.append(f"Row {row_num}: Nama Dealer kosong")
            return None

        # Parse tanggal masuk if provided (handle Excel numeric date format)
        tgl_masuk = tgl_datang
        if tgl_masuk_val:
            try:
                from datetime import datetime, timedelta
                if isinstance(tgl_masuk_val, (int, float)):
                    # Excel numeric date format (days since 1900-01-01)
                    # Excel starts from 1900-01-01 (with 1900 leap year bug)
                    excel_epoch = datetime(1899, 12, 30)
                    tgl_masuk = (excel_epoch + timedelta(days=int(tgl_masuk_val))).date()
                elif isinstance(tgl_masuk_val, str):
                    tgl_masuk = datetime.strptime(str(tgl_masuk_val), '%Y-%m-%d').date()
                elif isinstance(tgl_masuk_val, date):
                    tgl_masuk = tgl_masuk_val
            except Exception as e:
                # If parse fails, use default tgl_datang
                pass

        # Map dealer by name
        dealer_id = self._get_dealer_id_by_name(str(dealer_nama).strip())
        if not dealer_id:
            self.errors.append(f"Row {row_num}: Dealer '{dealer_nama}' tidak ditemukan")
            return None

        # Get or create type motor based on kode_type
        type_id = self._get_or_create_type_motor(str(kode_type).strip())
        if not type_id:
            self.errors.append(f"Row {row_num}: Gagal memproses Type '{kode_type}'")
            return None

        # Use warna value directly from Excel
        warna_value = str(warna).strip() if warna else ''

        # Parse Link/Ke Tujuan (optional - untuk terintegrasi pindah-stok)
        link_tujuan = values[6] if len(values) > 6 else None
        link_tujuan = str(link_tujuan).strip() if link_tujuan else None

        return {
            'row_num': row_num,
            'nama_dealer': dealer_nama,
            'no_mesin': str(no_mesin).strip(),
            'no_rangka': str(no_rangka).strip(),
            'type_id': type_id,
            'warna': warna_value,
            'dealer_id': dealer_id,
            'tgl_datang': tgl_masuk,
            'status': 'R',  # Ready (will be T if has Link/transfer)
            'link_tujuan': link_tujuan,
        }

    def _get_dealer_id_by_name(self, dealer_input: str) -> int:
        """Get dealer ID by name or kode (code)"""
        session = DatabaseManager.get_session()
        try:
            from database.models import Dealer

            # Mapping kode dealer singkat ke nama dealer
            dealer_mapping = {
                'JM1': 'Jaya Motor 1',
                'JM2': 'Jaya Motor 2',
                'JM3': 'Jaya Motor Bekasi',
                'A0035': 'Jaya Motor 1',
                'A0105': 'Jaya Motor 2',
                'A0210': 'Jaya Motor Bekasi',
            }

            # Try mapping first
            if dealer_input in dealer_mapping:
                dealer_input = dealer_mapping[dealer_input]

            # Try exact match by name
            dealer = session.query(Dealer).filter_by(nama=dealer_input).first()
            if dealer:
                return dealer.id

            # Try case-insensitive match by name
            dealer = session.query(Dealer).filter(
                Dealer.nama.ilike(dealer_input)
            ).first()
            if dealer:
                return dealer.id

            # Try match by kode
            dealer = session.query(Dealer).filter_by(kode_dealer=dealer_input).first()
            if dealer:
                return dealer.id

            return None
        except Exception as e:
            self.errors.append(f"Error looking up dealer '{dealer_input}': {str(e)}")
            return None
        finally:
            session.close()

    def _get_or_create_type_motor(self, kode_type: str) -> int:
        """
        Get TypeMotor ID if exists, or create new one if not exists

        Returns type_id or None if failed
        """
        session = DatabaseManager.get_session()
        try:
            # Check if type already exists
            existing = session.query(TypeMotor).filter_by(kode_type=kode_type).first()
            if existing:
                return existing.id

            # Create new type motor with basic info
            new_type = TypeMotor(
                kode_type=kode_type,
                nama_type=kode_type,  # Use kode as nama if not found
                status='A',
            )
            session.add(new_type)
            session.commit()
            type_id = new_type.id
            session.close()

            self.warnings.append(f"Type motor baru dibuat: {kode_type}")
            return type_id

        except Exception as e:
            self.errors.append(f"Error processing type motor '{kode_type}': {str(e)}")
            session.close()
            return None

    def _insert_to_database(self, rows: List[Dict]) -> int:
        """Insert validated rows to database and create StokTransfer if Link/Tujuan exists"""
        from database.models import Dealer, Broker, StokTransfer

        session = DatabaseManager.get_session()
        imported = 0

        try:
            for row_data in rows:
                # Check if already exists
                existing = session.query(StokMotor).filter_by(
                    no_mesin=row_data['no_mesin'],
                    no_rangka=row_data['no_rangka']
                ).first()

                if existing:
                    self.warnings.append(
                        f"Row {row_data['row_num']}: "
                        f"Unit {row_data['no_mesin']} sudah ada di database"
                    )
                    continue

                # Collect suggestions for type motor format
                type_id = row_data['type_id']
                if type_id not in self.type_motor_suggestions:
                    self.type_motor_suggestions[type_id] = {
                        'prefix_nomesin': row_data['no_mesin'][:5],      # First 5 digits
                        'prefix_norangka': row_data['no_rangka'][:7],    # First 7 digits
                    }

                # Create new stok motor
                stok = StokMotor(
                    no_mesin=row_data['no_mesin'],
                    no_rangka=row_data['no_rangka'],
                    type_id=row_data['type_id'],
                    warna=row_data['warna'],
                    dealer_id=row_data['dealer_id'],
                    tgl_datang=row_data['tgl_datang'],
                    status='R',  # Initial status: Ready
                )
                session.add(stok)
                session.flush()  # Flush to get stok.id

                # If Link/Tujuan provided, create StokTransfer and update motor status to T
                if row_data.get('link_tujuan'):
                    tujuan_dealer_id = None
                    tujuan_broker_id = None
                    tujuan_nama = row_data['link_tujuan']

                    # Try to find dealer by name (case-insensitive)
                    tujuan_dealer = session.query(Dealer).filter(
                        Dealer.nama.ilike(tujuan_nama)
                    ).first()
                    if tujuan_dealer:
                        tujuan_dealer_id = tujuan_dealer.id
                    else:
                        # Try to find broker by name (case-insensitive)
                        tujuan_broker = session.query(Broker).filter(
                            Broker.nama.ilike(tujuan_nama)
                        ).first()
                        if tujuan_broker:
                            tujuan_broker_id = tujuan_broker.id
                        else:
                            # Auto-create new dealer if not found
                            new_dealer = Dealer(
                                nama=tujuan_nama,
                                kode_dealer=f"AUTO_{tujuan_nama.upper()[:10]}",
                                status='A'
                            )
                            session.add(new_dealer)
                            session.flush()
                            tujuan_dealer_id = new_dealer.id
                            self.warnings.append(
                                f"Row {row_data['row_num']}: "
                                f"Dealer '{tujuan_nama}' dibuat otomatis"
                            )

                    if tujuan_dealer_id or tujuan_broker_id:
                        # Create transfer record
                        transfer = StokTransfer(
                            stok_motor_id=stok.id,
                            dealer_asal_id=row_data['dealer_id'],
                            dealer_tujuan_id=tujuan_dealer_id,
                            broker_tujuan_id=tujuan_broker_id,
                            tgl_transfer=row_data['tgl_datang'],
                            tipe_transfer='DISTRIBUSI',
                            status='A',  # Active transfer
                        )
                        session.add(transfer)

                        # Update motor status to Transferred
                        stok.status = 'T'

                imported += 1

            session.commit()
        except Exception as e:
            session.rollback()
            self.errors.append(f"Database error: {str(e)}")
        finally:
            session.close()

        return imported
